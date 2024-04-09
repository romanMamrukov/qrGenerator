import os
import openpyxl
import qrcode

def generate_qr_code(name, data, output_folder, filename_set):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )

    qr.add_data(data)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white", fit=True)

    img_name = f"{name}.png"

    if img_name in filename_set:

        return None  # QR code already generated for this name

    filename_set.add(img_name)

 

    img_path = os.path.join(output_folder, img_name)

    img.save(img_path)

    return img_path

 

def process_excel_file(excel_file_path, output_folder, filename_set):

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

 

    workbook = openpyxl.load_workbook(excel_file_path)

    sheet = workbook.active

 

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):

        name, data = [cell.value for cell in row]

        if data:
            img_path = generate_qr_code(name, data, output_folder, filename_set)

            if img_path:
                img = openpyxl.drawing.image.Image(img_path)
 

                cell_width = sheet.column_dimensions[row[1].column_letter].width
                cell_height = sheet.row_dimensions[row[1].row].height

 
                aspect_ratio = img.height / img.width
                img.width = cell_height / aspect_ratio
                img.height = cell_height

 
                img.anchor = row[1].offset(row=0, column=1).coordinate
                sheet.add_image(img)
 

    workbook.save(excel_file_path)
    print("QR codes generated and inserted into the Excel file.")


if __name__ == "__main__":

    excel_file_path = r"C:\Users\RomMam\Documents\0Projects\Codes\WebDev\2024\qrcode generator for excel\accountsQR.xlsx"
    output_folder = r"C:\Users\RomMam\Documents\0Projects\Codes\WebDev\2024\qrcode generator for excel\qr_codes"
    filename_set = set()  # Set to keep track of generated QR code filenames 

    process_excel_file(excel_file_path, output_folder, filename_set)